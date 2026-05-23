"""Общие утилиты для работы с Finances.xlsx.

Все скрипты проекта импортируют отсюда: WORKBOOK_PATH, BACKUP_DIR,
create_backup, atomic_save, assert_not_open, inventory.
"""
from __future__ import annotations

import datetime as _dt
import json
import os
import shutil
import subprocess
import sys
from pathlib import Path
from typing import Any

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
ROOT = PROJECT_ROOT  # для совместимости с импортирующими скриптами
WORKBOOK_PATH = PROJECT_ROOT / "data" / "legacy" / "Finances.xlsx"
BACKUP_DIR = PROJECT_ROOT / "data" / "legacy" / "backups"


def assert_not_open(path: Path) -> None:
    """Проверяет через lsof, что path не открыт другим процессом."""
    if not path.exists():
        return
    try:
        result = subprocess.run(
            ["lsof", str(path)], capture_output=True, text=True, timeout=5
        )
    except (subprocess.TimeoutExpired, FileNotFoundError):
        return
    if result.stdout.strip():
        sys.stderr.write(
            f"ERROR: {path.name} held by another process:\n{result.stdout}\n"
            "Close Excel/Numbers/LibreOffice and retry.\n"
        )
        sys.exit(2)


def create_backup(path: Path = WORKBOOK_PATH) -> Path:
    """Копирует path в backups/<basename>.<ts>.xlsx и возвращает путь к копии."""
    if not path.exists():
        sys.stderr.write(f"ERROR: source file not found: {path}\n")
        sys.exit(1)
    BACKUP_DIR.mkdir(exist_ok=True)
    ts = _dt.datetime.now().strftime("%Y-%m-%d_%H%M%S")
    dest = BACKUP_DIR / f"{path.stem}.{ts}{path.suffix}"
    shutil.copy2(path, dest)
    return dest


def atomic_save(wb, path: Path = WORKBOOK_PATH) -> None:
    """Сохраняет workbook через tmp-файл + os.replace для атомарности."""
    tmp = path.with_suffix(path.suffix + ".tmp")
    wb.save(tmp)
    os.replace(tmp, path)


def inventory(path: Path = WORKBOOK_PATH) -> dict[str, Any]:
    """Структурный инвентарь .xlsx — для diff и roundtrip-check."""
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=False)
    sheets: list[dict[str, Any]] = []
    total_formulas = 0
    for ws in wb.worksheets:
        formulas = sum(
            1
            for row in ws.iter_rows(values_only=False)
            for cell in row
            if cell.data_type == "f"
        )
        total_formulas += formulas
        cf_count = 0
        try:
            cf_count = sum(
                len(v) for v in ws.conditional_formatting._cf_rules.values()
            )
        except Exception:
            pass
        sheets.append(
            {
                "name": ws.title,
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "charts": len(ws._charts),
                "images": len(ws._images),
                "tables": len(ws.tables),
                "data_validations": len(ws.data_validations.dataValidation),
                "conditional_formats": cf_count,
                "merged_cells": len(ws.merged_cells.ranges),
                "freeze_panes": ws.freeze_panes,
                "formulas": formulas,
                "protected": bool(ws.protection.sheet),
            }
        )
    return {
        "path": str(path),
        "size_bytes": path.stat().st_size,
        "mtime": _dt.datetime.fromtimestamp(path.stat().st_mtime).isoformat(),
        "sheets": sheets,
        "total_formulas": total_formulas,
        "defined_names": list(wb.defined_names),
        "vba": getattr(wb, "vba_archive", None) is not None,
    }


def print_json(obj: Any) -> None:
    print(json.dumps(obj, indent=2, ensure_ascii=False, default=str))
