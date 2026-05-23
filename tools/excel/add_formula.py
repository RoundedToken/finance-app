#!/usr/bin/env python3
"""add_formula.py — записывает формулу в указанную ячейку.

Делает backup и atomic save. Не пересчитывает значения — для пересчёта
запустить `python scripts/calc.py`.

Usage:
    python scripts/add_formula.py Sheet1 C10 "=SUM(C2:C9)"
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.append(str(Path(__file__).resolve().parent))
from _common import (  # noqa: E402
    WORKBOOK_PATH,
    assert_not_open,
    atomic_save,
    create_backup,
    inventory,
)


def _structure_diff(before: dict, after: dict) -> list[str]:
    diffs: list[str] = []
    for sa, sb in zip(before["sheets"], after["sheets"]):
        for k in ("charts", "images", "tables", "data_validations",
                  "conditional_formats", "merged_cells"):
            if sa[k] != sb[k]:
                diffs.append(f"{sa['name']}.{k}: {sa[k]} → {sb[k]}")
    return diffs


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("sheet")
    ap.add_argument("address")
    ap.add_argument("formula", help="must start with '='")
    ap.add_argument("--file", default=str(WORKBOOK_PATH), type=Path)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    if not args.formula.startswith("="):
        sys.stderr.write("ERROR: formula must start with '='\n")
        return 1

    assert_not_open(args.file)
    from openpyxl import load_workbook

    wb = load_workbook(args.file)
    ws = wb[args.sheet]
    print(f"{args.sheet}!{args.address}: {ws[args.address].value!r} → {args.formula}")
    if args.dry_run:
        return 0

    before = inventory(args.file)
    backup = create_backup(args.file)
    print(f"backup: {backup}")
    ws[args.address] = args.formula
    atomic_save(wb, args.file)
    after = inventory(args.file)

    diffs = _structure_diff(before, after)
    if diffs:
        sys.stderr.write(
            "WARNING — structure changed:\n  " + "\n  ".join(diffs) + "\n"
        )
        sys.stderr.write(f"To revert: python scripts/restore.py {backup}\n")
        return 3
    print("OK — formula saved. Run `python scripts/calc.py` to refresh cached values.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
