#!/usr/bin/env python3
"""set_cell.py — атомарно записывает значение в одну ячейку.

Делает backup до записи, atomic save после, и сверяет инвентарь (charts,
tables, named ranges) — не должны исчезнуть.

Usage:
    python scripts/set_cell.py <sheet> <address> <value>
    python scripts/set_cell.py Sheet1 B5 "1234.56"
    python scripts/set_cell.py --dry-run Sheet1 B5 "new"

Значения парсятся: int / float / true / false → bool / "" → None / иначе строка.
Для формул использовать add_formula.py.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.append(str(Path(__file__).resolve().parent))
from _common import (  # noqa: E402
    WORKBOOK_PATH,
    assert_not_open,
    atomic_save,
    create_backup,
    inventory,
)


def parse_value(s: str):
    sl = s.strip().lower()
    if sl in {"true", "false"}:
        return sl == "true"
    if sl in {"none", "null", ""}:
        return None
    try:
        if "." in s or "e" in sl:
            return float(s)
        return int(s)
    except ValueError:
        return s


def _structure_diff(before: dict, after: dict) -> list[str]:
    diffs: list[str] = []
    for sa, sb in zip(before["sheets"], after["sheets"]):
        for k in ("charts", "images", "tables", "data_validations",
                  "conditional_formats", "merged_cells"):
            if sa[k] != sb[k]:
                diffs.append(f"{sa['name']}.{k}: {sa[k]} → {sb[k]}")
    if len(before["defined_names"]) != len(after["defined_names"]):
        diffs.append(
            f"defined_names: {len(before['defined_names'])} → {len(after['defined_names'])}"
        )
    return diffs


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("sheet")
    ap.add_argument("address", help="cell address like B5")
    ap.add_argument("value")
    ap.add_argument("--file", default=str(WORKBOOK_PATH), type=Path)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    assert_not_open(args.file)
    from openpyxl import load_workbook

    wb = load_workbook(args.file)
    ws = wb[args.sheet]
    new_value = parse_value(args.value)
    old_value = ws[args.address].value
    print(f"{args.sheet}!{args.address}: {old_value!r} → {new_value!r}")
    if args.dry_run:
        print("dry-run: not saving")
        return 0

    before = inventory(args.file)
    backup = create_backup(args.file)
    print(f"backup: {backup}")
    ws[args.address] = new_value
    atomic_save(wb, args.file)
    after = inventory(args.file)

    diffs = _structure_diff(before, after)
    if diffs:
        sys.stderr.write(
            "WARNING — structure changed:\n  " + "\n  ".join(diffs) + "\n"
        )
        sys.stderr.write(f"To revert: python scripts/restore.py {backup}\n")
        return 3
    print("OK — saved, inventory intact")
    return 0


if __name__ == "__main__":
    sys.exit(main())
