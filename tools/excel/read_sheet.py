#!/usr/bin/env python3
"""read_sheet.py — выводит содержимое листа в CSV или JSON.

Usage:
    python scripts/read_sheet.py                                # первый лист, CSV
    python scripts/read_sheet.py --sheet "Sheet1"
    python scripts/read_sheet.py --sheet 0 --range A1:E10
    python scripts/read_sheet.py --values-only                  # cached values
    python scripts/read_sheet.py --json
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path

sys.path.append(str(Path(__file__).resolve().parent))
from _common import WORKBOOK_PATH  # noqa: E402


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--file", default=str(WORKBOOK_PATH), type=Path)
    ap.add_argument("--sheet", default="0", help="name or 0-based index")
    ap.add_argument("--range", dest="cell_range", default=None)
    ap.add_argument(
        "--values-only",
        action="store_true",
        help="cached values instead of formula source",
    )
    ap.add_argument("--json", action="store_true")
    args = ap.parse_args()

    from openpyxl import load_workbook

    wb = load_workbook(args.file, data_only=args.values_only)
    try:
        idx = int(args.sheet)
        ws = wb.worksheets[idx]
    except ValueError:
        ws = wb[args.sheet]

    rows = list(ws[args.cell_range]) if args.cell_range else list(ws.iter_rows())
    data = [[c.value for c in row] for row in rows]

    if args.json:
        json.dump(
            {"sheet": ws.title, "range": args.cell_range, "rows": data},
            sys.stdout,
            ensure_ascii=False,
            indent=2,
            default=str,
        )
    else:
        w = csv.writer(sys.stdout)
        for r in data:
            w.writerow(["" if v is None else v for v in r])
    return 0


if __name__ == "__main__":
    sys.exit(main())
