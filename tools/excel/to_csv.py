#!/usr/bin/env python3
"""to_csv.py — экспорт листов в CSV.

Usage:
    python scripts/to_csv.py                                # все листы в csv/
    python scripts/to_csv.py --sheet "Sheet1" -o sheet1.csv
"""
from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

sys.path.append(str(Path(__file__).resolve().parent))
from _common import ROOT, WORKBOOK_PATH  # noqa: E402


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--file", default=str(WORKBOOK_PATH), type=Path)
    ap.add_argument("--sheet", default=None)
    ap.add_argument("--with-formulas", action="store_true",
                    help="write formula source instead of cached values")
    ap.add_argument("-o", "--out", default=None, type=Path)
    args = ap.parse_args()

    from openpyxl import load_workbook

    wb = load_workbook(args.file, data_only=not args.with_formulas)
    sheets = [args.sheet] if args.sheet else wb.sheetnames

    if args.out and args.out.suffix == ".csv":
        outdir = args.out.parent
        single_path: Path | None = args.out
    else:
        outdir = args.out or (ROOT / "csv")
        single_path = None
    outdir.mkdir(parents=True, exist_ok=True)

    for name in sheets:
        ws = wb[name]
        target = single_path if single_path else outdir / f"{name}.csv"
        with open(target, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                w.writerow(["" if v is None else v for v in row])
        print(f"wrote: {target}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
