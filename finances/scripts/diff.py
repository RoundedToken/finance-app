#!/usr/bin/env python3
"""diff.py — структурный diff двух .xlsx (инвентарь + опционально значения).

Usage:
    python scripts/diff.py Finances.xlsx backups/Finances.<ts>.xlsx
    python scripts/diff.py --show-cells Finances.xlsx backups/Finances.<ts>.xlsx
    python scripts/diff.py --sheet Sheet1 --show-cells a.xlsx b.xlsx
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _common import inventory  # noqa: E402


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("a", type=Path)
    ap.add_argument("b", type=Path)
    ap.add_argument("--sheet", default=None)
    ap.add_argument(
        "--show-cells",
        action="store_true",
        help="also list changed cells (slower, scans every cell)",
    )
    args = ap.parse_args()

    ia, ib = inventory(args.a), inventory(args.b)
    print(f"--- {args.a}")
    print(f"+++ {args.b}")
    print(f"size:           {ia['size_bytes']} → {ib['size_bytes']}")
    print(f"sheets:         {[s['name'] for s in ia['sheets']]} → {[s['name'] for s in ib['sheets']]}")
    print(f"defined_names:  {ia['defined_names']} → {ib['defined_names']}")
    print(f"total_formulas: {ia['total_formulas']} → {ib['total_formulas']}")

    a_sheets = {s["name"]: s for s in ia["sheets"]}
    b_sheets = {s["name"]: s for s in ib["sheets"]}
    for name in sorted(set(a_sheets) | set(b_sheets)):
        sa, sb = a_sheets.get(name), b_sheets.get(name)
        if not (sa and sb):
            print(f"sheet {name}: present in only one file")
            continue
        for k in ("max_row", "max_col", "formulas", "charts", "tables",
                  "data_validations", "conditional_formats", "merged_cells"):
            if sa[k] != sb[k]:
                print(f"  {name}.{k}: {sa[k]} → {sb[k]}")

    if args.show_cells:
        from openpyxl import load_workbook

        wa = load_workbook(args.a)
        wb_ = load_workbook(args.b)
        sheets = (
            [args.sheet]
            if args.sheet
            else [s for s in wa.sheetnames if s in wb_.sheetnames]
        )
        for name in sheets:
            sa = wa[name]
            sb = wb_[name]
            rows = max(sa.max_row, sb.max_row)
            cols = max(sa.max_column, sb.max_column)
            for row in range(1, rows + 1):
                for col in range(1, cols + 1):
                    va = sa.cell(row, col).value
                    vb = sb.cell(row, col).value
                    if va != vb:
                        addr = sa.cell(row, col).coordinate
                        print(f"  {name}!{addr}: {va!r} → {vb!r}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
